#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Sistema de Portfolio Trading Neutro - FuzzyFajuto

Versão: 2.0.0 (Simplified)
"""

import pandas as pd
import numpy as np
import os
import json
import logging
import datetime
import traceback
import yfinance as yf
import talib
from openpyxl import load_workbook
import sys
import warnings
import locale
from typing import Dict, List, Tuple, Optional

# Import configuration from config.py
from config import get_config, ensure_directories, CONFIG, DATA_DIR, RESULTS_DIR

# Configuração de logs
def setup_logging():
    """Configura o sistema de logs"""
    from config import LOG_DIR
    
    ensure_directories()
    log_file = os.path.join(LOG_DIR, f"fuzzyfajuto_{datetime.datetime.now().strftime('%Y%m%d')}.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    
    return logging.getLogger("FuzzyFajuto")

# Logger global
logger = setup_logging()

# Obter configuração
CONFIG = get_config()

# Adicionar o diretório raiz ao path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Desabilitar warnings
warnings.filterwarnings('ignore')

def truncate_price(price, decimals=2):
    """
    Trunca o preço para o número especificado de casas decimais
    """
    if price is None:
        return None
    factor = 10 ** decimals
    return int(price * factor) / factor

def get_stock_data(symbols, start_date, end_date):
    """
    Obtém dados históricos para uma lista de ações
    
    Args:
        symbols (list): Lista de símbolos para obter dados
        start_date (datetime): Data inicial
        end_date (datetime): Data final
        
    Returns:
        dict: Dicionário com dados históricos para cada símbolo
    """
    data = {}
    ibov_symbol = CONFIG["market"]["ibov_symbol"]
    
    # Adiciona o Ibovespa à lista
    all_symbols = symbols.copy()
    if ibov_symbol not in all_symbols:
        all_symbols.append(ibov_symbol)
    
    logger.info(f"Obtendo dados para {len(all_symbols)} ativos")
    
    for symbol in all_symbols:
        try:
            logger.info(f"Baixando dados de {symbol}")
            stock_data = yf.download(symbol, start=start_date, end=end_date, progress=False)
            
            if not stock_data.empty:
                # Verificar e tratar MultiIndex em colunas que o Yahoo Finance retorna
                if isinstance(stock_data.columns, pd.MultiIndex):
                    # Simplificar os dados transformando o MultiIndex em colunas simples
                    flattened_data = pd.DataFrame()
                    
                    # Verificar se temos um nível de ticker
                    if len(stock_data.columns.levels) >= 2:
                        ticker_level = 1  # Geralmente o ticker está no nível 1
                        
                        if symbol in stock_data.columns.levels[ticker_level]:
                            # Selecionar apenas os dados para este símbolo
                            flattened_data = stock_data.xs(symbol, axis=1, level=ticker_level)
                        else:
                            # Se não encontrar o ticker específico, aplainar o MultiIndex
                            for col in stock_data.columns:
                                flattened_data[col[0]] = stock_data[col]
                    else:
                        # Aplainar o MultiIndex se não tiver nível de ticker
                        for col in stock_data.columns:
                            flattened_data[col[0]] = stock_data[col]
                    
                    data[symbol] = flattened_data
                else:
                    # Já é um DataFrame regular
                    data[symbol] = stock_data
            else:
                logger.warning(f"Dados vazios para {symbol}")
        except Exception as e:
            logger.error(f"Erro ao obter dados para {symbol}: {e}")
    
    return data

def calculate_indicators(stock_data):
    """
    Calcula indicadores técnicos para os dados de ações
    
    Args:
        stock_data (dict): Dicionário com dados históricos
        
    Returns:
        dict: Dados com indicadores calculados
    """
    ema_periods = CONFIG["indicators"]["ema_periods"]
    rsi_period = CONFIG["indicators"]["rsi_period"]
    
    for symbol, data in stock_data.items():
        if not data.empty and len(data) > max(ema_periods + [rsi_period]):
            try:
                # Verificar se os dados têm formato válido
                if 'Close' not in data.columns:
                    logger.warning(f"Coluna 'Close' não encontrada para {symbol}, pulando...")
                    continue
                    
                close_values = data['Close'].values
                if len(close_values) < 2:
                    logger.warning(f"Dados insuficientes para {symbol}, pulando...")
                    continue
                
                # Convertendo para floats para garantir compatibilidade com TA-Lib
                close_values = close_values.astype(float)
                
                # Calculando EMAs para cada período configurado
                for period in ema_periods:
                    if len(close_values) > period:
                        data[f'EMA_{period}'] = talib.EMA(close_values, timeperiod=period)
                    else:
                        logger.warning(f"Período EMA {period} maior que os dados disponíveis para {symbol}")
                        data[f'EMA_{period}'] = float('nan')
                
                # Calculando RSI
                if len(close_values) > rsi_period:
                    data[f'RSI_{rsi_period}'] = talib.RSI(close_values, timeperiod=rsi_period)
                else:
                    logger.warning(f"Período RSI {rsi_period} maior que os dados disponíveis para {symbol}")
                    data[f'RSI_{rsi_period}'] = float('nan')
                
                # Calculando retornos
                data['Return'] = data['Close'].pct_change()
            except Exception as e:
                logger.error(f"Erro ao calcular indicadores para {symbol}: {e}")
    
    return stock_data

def calculate_fuzzy_fajuto_score(stock_data, date):
    """
    Calcula o score FuzzyFajuto para todas as ações em uma data específica
    
    Args:
        stock_data (dict): Dicionário com dados históricos e indicadores
        date (datetime): Data para calcular o score
        
    Returns:
        dict: Scores calculados para cada ativo
    """
    scores = {}
    ibov_symbol = CONFIG["market"]["ibov_symbol"]
    ema_periods = CONFIG["indicators"]["ema_periods"]
    rsi_period = CONFIG["indicators"]["rsi_period"]
    overbought = CONFIG["indicators"]["overbought_threshold"]
    oversold = CONFIG["indicators"]["oversold_threshold"]
    
    # Verificar se temos dados para o Ibovespa
    if ibov_symbol not in stock_data or date not in stock_data[ibov_symbol].index:
        logger.error(f"Dados do Ibovespa não disponíveis para {date}")
        if ibov_symbol in stock_data:
            # Tenta usar a data mais recente disponível
            last_date = stock_data[ibov_symbol].index[-1]
            logger.warning(f"Usando a data mais recente disponível: {last_date}")
            date = last_date
        else:
            logger.error("Ibovespa não encontrado nos dados, impossível calcular scores")
            return {}
    
    # Obter retorno do IBOV
    ibov_return = stock_data[ibov_symbol].loc[date, 'Return'] if date in stock_data[ibov_symbol].index else 0
    if pd.isna(ibov_return):
        ibov_return = 0
        logger.warning(f"Retorno do Ibovespa para {date} é NaN, usando 0")
    
    for symbol, data in stock_data.items():
        if symbol == ibov_symbol:
            continue
            
        if date not in data.index:
            logger.debug(f"Data {date} não disponível para {symbol}, pulando...")
            continue
            
        try:
            # Verificar se temos todos os indicadores necessários
            required_columns = ['Close', 'Return'] + [f'EMA_{p}' for p in ema_periods] + [f'RSI_{rsi_period}']
            missing_columns = [col for col in required_columns if col not in data.columns]
            
            if missing_columns:
                logger.debug(f"Colunas ausentes para {symbol}: {missing_columns}, pulando...")
                continue
            
            # Obter dados do dia
            close = data.loc[date, 'Close']
            stock_return = data.loc[date, 'Return']
            
            # Lidar com NaN no retorno
            if pd.isna(stock_return):
                logger.debug(f"Retorno para {symbol} é NaN, usando 0")
                stock_return = 0
            
            # Component 1: Retorno do ativo vs retorno do IBOV
            component1 = 0
            if stock_return > ibov_return and stock_return > 0:
                component1 = 1
            elif stock_return < ibov_return and stock_return < 0:
                component1 = -1
                
            # Component 2: Fechamento vs EMAs (média de todas as comparações)
            component2 = 0
            ema_count = 0
            
            for period in ema_periods:
                ema_value = data.loc[date, f'EMA_{period}']
                if not pd.isna(ema_value):
                    component2 += 1 if close > ema_value else -1
                    ema_count += 1
                
            if ema_count > 0:
                component2 = component2 / ema_count
            else:
                logger.debug(f"Nenhum EMA válido para {symbol}, usando 0 para component2")
            
            # Component 3: RSI
            component3 = 0
            rsi_value = data.loc[date, f'RSI_{rsi_period}']
            
            if not pd.isna(rsi_value):
                if rsi_value > overbought:
                    component3 = -0.75  # Sobrecomprado
                elif rsi_value < oversold:
                    component3 = 0.75   # Sobrevendido
            else:
                logger.debug(f"RSI para {symbol} é NaN, usando 0 para component3")
            
            # Score final (valores entre -2.75 e +2.75)
            score = component1 + component2 + component3
            scores[symbol] = {
                'score': score,
                'close': close,
                'rsi': rsi_value if not pd.isna(rsi_value) else None,
                'return': stock_return,
                'vs_ibov': stock_return - ibov_return
            }
            
            logger.debug(f"Score calculado para {symbol}: {score:.2f}")
        except Exception as e:
            logger.error(f"Erro ao calcular score para {symbol}: {e}")
    
    return scores

def select_portfolio(scores):
    """
    Seleciona os ativos com base nos maiores scores absolutos FuzzyFajuto
    e ordena por preço para equilibrar em mais ativos
    
    Args:
        scores (dict): Dicionário com scores calculados
        
    Returns:
        dict: Portfólio selecionado para compra e venda
    """
    max_positions = CONFIG["model"]["max_position_per_side"]
    
    # Separar em compras e vendas, mas priorizar por magnitude do score
    buy_candidates = []
    sell_candidates = []
    
    # Encontrar os melhores scores para compra e venda
    best_buy_score = 0
    best_sell_score = 0
    
    for symbol, data in scores.items():
        score = data['score']
        
        # Atualizar os melhores scores
        if score > best_buy_score:
            best_buy_score = score
        elif score < best_sell_score:
            best_sell_score = score
    
    # Definir thresholds para filtrar candidatos com base nos melhores scores
    buy_threshold = best_buy_score - 0.5
    sell_threshold = best_sell_score + 0.5  # Inverte para scores negativos
    
    logger.info(f"Melhor score de compra: {best_buy_score:.2f}, threshold: {buy_threshold:.2f}")
    logger.info(f"Melhor score de venda: {best_sell_score:.2f}, threshold: {sell_threshold:.2f}")
    
    for symbol, data in scores.items():
        score = data['score']
        close = data['close']
        
        # Adicionar a todos os candidatos com informações de preço
        if score > 0 and score >= buy_threshold:
            buy_candidates.append({
                'symbol': symbol,
                'score': score,
                'close': close,
                'abs_score': abs(score)  # Para ordenação por magnitude
            })
        elif score < 0 and score <= sell_threshold:
            sell_candidates.append({
                'symbol': symbol,
                'score': score,
                'close': close,
                'abs_score': abs(score)  # Para ordenação por magnitude
            })
    
    # Ordenar candidatos primeiro por magnitude do score (maior primeiro)
    # e depois por preço (menor primeiro) para diversificar
    buy_candidates.sort(key=lambda x: (-x['abs_score'], x['close']))
    sell_candidates.sort(key=lambda x: (-x['abs_score'], x['close']))
    
    # Limitar ao número máximo de posições
    logger.info(f"Candidatos à compra (após filtro de threshold): {len(buy_candidates)}")
    logger.info(f"Candidatos à venda (após filtro de threshold): {len(sell_candidates)}")
    logger.info(f"Selecionando até {max_positions} posições de cada lado")
    
    return {
        'buys': buy_candidates[:max_positions],
        'sells': sell_candidates[:max_positions]
    }

def filter_portfolio_by_quality(portfolio):
    """
    Filtra o portfólio para manter apenas posições com scores acima de 1.50 (compra)
    e abaixo de -1.50 (venda).
    
    Args:
        portfolio (dict): Portfólio com posições de compra e venda
        
    Returns:
        dict: Portfólio filtrado por qualidade
    """
    # Definir limites fixos de qualidade
    buy_quality_threshold = 1.50
    sell_quality_threshold = -1.50
    
    logger.info(f"Score mínimo de compra: >= {buy_quality_threshold:.2f}")
    logger.info(f"Score máximo de venda: <= {sell_quality_threshold:.2f}")
    
    # Filtrar posições
    quality_buys = [item for item in portfolio['buys'] if item['score'] >= buy_quality_threshold]
    quality_sells = [item for item in portfolio['sells'] if item['score'] <= sell_quality_threshold]
    
    # Registrar quantidade de posições após filtro
    logger.info(f"Após filtro por qualidade - Compra: {len(quality_buys)}/{len(portfolio['buys'])}, Venda: {len(quality_sells)}/{len(portfolio['sells'])}")
    
    return {
        'buys': quality_buys,
        'sells': quality_sells
    }

def generate_tryd_orders(portfolio, exposicao_financeira_total_param=None):
    """
    Gera as ordens para o Tryd Automate, distribuindo a exposição financeira desejada
    de forma igual entre 4 ordens por ativo conforme TRADING_LOGIC.md
    
    Args:
        portfolio (dict): Portfólio selecionado
        exposicao_financeira_total_param (float, optional): Exposição financeira total desejada (R$). 
                                                        Se None, usa o valor da configuração.
        
    Returns:
        dict: Ordens geradas
    """
    # Função auxiliar para encontrar o score original de um ativo
    def get_original_score(symbol, candidates):
        for candidate in candidates:
            if candidate['symbol'].replace('.SA', '') == symbol:
                return candidate['score']
        return 0
    
    # Filtrar o portfólio por qualidade de score
    filtered_portfolio = filter_portfolio_by_quality(portfolio)
    
    # Verificar se temos o score de pelo menos 1º e 2º graus
    buy_scores = [item['score'] for item in filtered_portfolio['buys']] if filtered_portfolio['buys'] else []
    sell_scores = [item['score'] for item in filtered_portfolio['sells']] if filtered_portfolio['sells'] else []
    
    if buy_scores:
        unique_buy_scores = sorted(set(buy_scores), reverse=True)
        buy_grades = len(unique_buy_scores)
        logger.info(f"Scores de compra: {unique_buy_scores[:3]} - {buy_grades} graus diferentes")
    
    if sell_scores:
        unique_sell_scores = sorted(set(sell_scores), reverse=False)
        sell_grades = len(unique_sell_scores)
        logger.info(f"Scores de venda: {unique_sell_scores[:3]} - {sell_grades} graus diferentes")
    
    # Usar o portfólio filtrado para o restante da função
    portfolio = filtered_portfolio
    
    # Carregando configurações específicas
    max_exposure_per_stock = CONFIG["model"].get("max_exposure_per_stock", 0.15)  # 15% de exposição por ativo
    lot_size_expensive = CONFIG["execution"].get("lot_size_expensive", 10)  # Lote para ações > R$50
    lot_size_normal = CONFIG["execution"].get("lot_size_normal", 100)     # Lote para ações normais
    
    # Definir 4 níveis de preço para ordens conforme TRADING_LOGIC.md
    # Para compras: mercado, -0.5%, -1.0%, -1.5%
    # Para vendas: mercado, +0.5%, +1.0%, +1.5%
    price_levels_buy = [
        {"type": "MARKET", "adjustment": 0.0, "description": "Mercado"},
        {"type": "LIMIT", "adjustment": -0.005, "description": "-0.5%"},
        {"type": "LIMIT", "adjustment": -0.01, "description": "-1.0%"},
        {"type": "LIMIT", "adjustment": -0.015, "description": "-1.5%"}
    ]
    
    price_levels_sell = [
        {"type": "MARKET", "adjustment": 0.0, "description": "Mercado"},
        {"type": "LIMIT", "adjustment": 0.005, "description": "+0.5%"},
        {"type": "LIMIT", "adjustment": 0.01, "description": "+1.0%"},
        {"type": "LIMIT", "adjustment": 0.015, "description": "+1.5%"}
    ]
    
    # Determinar a exposição financeira final
    if exposicao_financeira_total_param is None:
        exposicao_financeira_final = CONFIG["model"]["exposicao_financeira_total"]
    else:
        exposicao_financeira_final = exposicao_financeira_total_param
    
    # Pegar o número máximo de posições por lado da configuração
    max_positions_per_side = CONFIG["model"]["max_position_per_side"]
    
    # Calcular exposição por lado
    exposicao_por_lado = exposicao_financeira_final / 2.0
    
    # Calcular o limite máximo de exposição por ativo (% do lado)
    max_exposicao_por_ativo = exposicao_por_lado * max_exposure_per_stock
    logger.info(f"Limite de exposição por ativo: R$ {max_exposicao_por_ativo:.2f} ({max_exposure_per_stock*100}% de R$ {exposicao_por_lado:.2f})")
    
    # Calcular a exposição inicial por ativo (podemos precisar ajustar isso depois)
    num_buys = min(len(portfolio['buys']), max_positions_per_side)
    num_sells = min(len(portfolio['sells']), max_positions_per_side)
    
    # Exposição por posição (não por ativo)
    # Conforme TRADING_LOGIC.md: EXPOSURE_PER_POSITION = 5000
    exposure_per_position = 15000.0
    
    # Estruturas para armazenar ordens e ativos alocados
    buy_stocks_used = []
    sell_stocks_used = []
    all_buy_orders = []
    all_sell_orders = []
    
    # FASE 1: Alocação de todas as posições possíveis em compra e venda
    # Alocação lado compra
    exposicao_acumulada_compra = 0.0
    
    for buy in portfolio['buys'][:max_positions_per_side]:
        symbol = buy['symbol'].replace('.SA', '')
        close = buy['close']
        score = buy['score']
        
        # Quantidade mínima baseada no preço (lote padrão)
        min_qty = lot_size_expensive if close > 50 else lot_size_normal
        
        # Calcular exposição para este ativo (fixa em 5000 conforme TRADING_LOGIC.md)
        target_exposure = exposure_per_position
        
        # Verificar orçamento disponível
        if exposicao_acumulada_compra + target_exposure > exposicao_por_lado:
            continue
            
        # Calcular quantidade total necessária para exposição de R$ 5000
        total_shares_needed = target_exposure / close
        
        # Arredondar para o lote padrão
        lot_size = lot_size_expensive if close > 50 else lot_size_normal
        total_qty = int(total_shares_needed // lot_size) * lot_size
        
        # Verificar quantidade mínima
        if total_qty < min_qty:
            continue
        
        # Calcular exposição real
        actual_exposure = total_qty * close
        
        # Atualizar exposição acumulada
        exposicao_acumulada_compra += actual_exposure
        
        # Adicionar à lista de ações usadas
        buy_stocks_used.append({
            'symbol': symbol,
            'qty': total_qty,
            'price': close,
            'exposure': actual_exposure,
            'score': score
        })
        
        # Dividir a quantidade total em 4 partes iguais para as 4 ordens
        # Conforme TRADING_LOGIC.md: raw_quantity = (exposure / 4) / price
        qty_per_order = total_qty // 4
        
        # Garantir que cada ordem tenha pelo menos o lote mínimo
        if qty_per_order < lot_size:
            qty_per_order = lot_size
        
        # Ajustar quantidades para lotes
        qty_per_order = int(qty_per_order // lot_size) * lot_size
        
        # Se a divisão não for exata, adicionar o resto à primeira ordem
        remainder = total_qty - (qty_per_order * 4)
        
        # Gerar 4 ordens com quantidades iguais
        for i, level in enumerate(price_levels_buy):
            # Adicionar o resto à primeira ordem se houver
            order_qty = qty_per_order + (remainder if i == 0 else 0)
            
            if order_qty > 0:
                price_adjustment = level["adjustment"]
                order_type = level["type"]
                
                if order_type == "MARKET":
                    # Ordem a mercado
                    price_adjusted = close
                    price_display = "MARKET"
                else:
                    # Ordem limitada - usar truncamento conforme TRADING_LOGIC.md
                    price_adjusted = truncate_price(close * (1 + price_adjustment), 2)
                    price_display = price_adjusted
                
                level_exposure = order_qty * close
                
                # Adicionar ordem
                all_buy_orders.append({
                    'stock': symbol,
                    'action': 'BUY',
                    'qty': order_qty,
                    'price': price_display,
                    'actual_price': price_adjusted,
                    'type': order_type,
                    'score': score,
                    'exposure': level_exposure,
                    'description': level["description"]
                })
    
    # Alocação lado venda
    exposicao_acumulada_venda = 0.0
    
    for sell in portfolio['sells'][:max_positions_per_side]:
        symbol = sell['symbol'].replace('.SA', '')
        close = sell['close']
        score = sell['score']
        
        # Quantidade mínima baseada no preço (lote padrão)
        min_qty = lot_size_expensive if close > 50 else lot_size_normal
        
        # Calcular exposição para este ativo (fixa em 5000 conforme TRADING_LOGIC.md)
        target_exposure = exposure_per_position
        
        # Verificar orçamento disponível
        if exposicao_acumulada_venda + target_exposure > exposicao_por_lado:
            continue
            
        # Calcular quantidade total necessária para exposição de R$ 5000
        total_shares_needed = target_exposure / close
        
        # Arredondar para o lote padrão
        lot_size = lot_size_expensive if close > 50 else lot_size_normal
        total_qty = int(total_shares_needed // lot_size) * lot_size
        
        # Verificar quantidade mínima
        if total_qty < min_qty:
            continue
        
        # Calcular exposição real
        actual_exposure = total_qty * close
        
        # Atualizar exposição acumulada
        exposicao_acumulada_venda += actual_exposure
        
        # Adicionar à lista de ações usadas
        sell_stocks_used.append({
            'symbol': symbol,
            'qty': total_qty,
            'price': close,
            'exposure': actual_exposure,
            'score': score
        })
        
        # Dividir a quantidade total em 4 partes iguais para as 4 ordens
        # Conforme TRADING_LOGIC.md: raw_quantity = (exposure / 4) / price
        qty_per_order = total_qty // 4
        
        # Garantir que cada ordem tenha pelo menos o lote mínimo
        if qty_per_order < lot_size:
            qty_per_order = lot_size
        
        # Ajustar quantidades para lotes
        qty_per_order = int(qty_per_order // lot_size) * lot_size
        
        # Se a divisão não for exata, adicionar o resto à primeira ordem
        remainder = total_qty - (qty_per_order * 4)
        
        # Gerar 4 ordens com quantidades iguais
        for i, level in enumerate(price_levels_sell):
            # Adicionar o resto à primeira ordem se houver
            order_qty = qty_per_order + (remainder if i == 0 else 0)
            
            if order_qty > 0:
                price_adjustment = level["adjustment"]
                order_type = level["type"]
                
                if order_type == "MARKET":
                    # Ordem a mercado
                    price_adjusted = close
                    price_display = "MARKET"
                else:
                    # Ordem limitada - usar truncamento conforme TRADING_LOGIC.md
                    price_adjusted = truncate_price(close * (1 + price_adjustment), 2)
                    price_display = price_adjusted
                
                level_exposure = order_qty * close
                
                # Adicionar ordem
                all_sell_orders.append({
                    'stock': symbol,
                    'action': 'SELL',
                    'qty': order_qty,
                    'price': price_display,
                    'actual_price': price_adjusted,
                    'type': order_type,
                    'score': score,
                    'exposure': level_exposure,
                    'description': level["description"]
                })
    
    # FASE 2: Balancear o número final de posições após alocação
    buy_stocks_count = len(buy_stocks_used)
    sell_stocks_count = len(sell_stocks_used)
    
    # Balancear para o menor número de posições
    min_positions = min(buy_stocks_count, sell_stocks_count)
    
    if min_positions == 0:
        logger.warning("Não foi possível alocar posições suficientes em um dos lados. Verifique as restrições de orçamento e lote mínimo.")
        min_positions = max(buy_stocks_count, sell_stocks_count)  # Usar o lado que tem posições
    
    # Ordenar posições por score (decrescente) e depois por exposição (decrescente)
    buy_stocks_used.sort(key=lambda x: (-abs(x['score']), -x['exposure']))
    sell_stocks_used.sort(key=lambda x: (-abs(x['score']), -x['exposure']))
    
    # Selecionar as melhores posições de cada lado
    balanced_buys = buy_stocks_used[:min_positions]
    balanced_sells = sell_stocks_used[:min_positions]
    
    logger.info(f"Posições após balanceamento final: {len(balanced_buys)} compras e {len(balanced_sells)} vendas")
    
    # Filtrar ordens apenas para os ativos selecionados após balanceamento
    buy_symbols = {stock['symbol'] for stock in balanced_buys}
    sell_symbols = {stock['symbol'] for stock in balanced_sells}
    
    balanced_buy_orders = [order for order in all_buy_orders if order['stock'] in buy_symbols]
    balanced_sell_orders = [order for order in all_sell_orders if order['stock'] in sell_symbols]
    
    # Calcular exposição final após balanceamento
    exposicao_final_compra = sum(stock['exposure'] for stock in balanced_buys)
    exposicao_final_venda = sum(stock['exposure'] for stock in balanced_sells)
    
    # Retornar ordens balanceadas
    orders = {
        'buys': balanced_buy_orders,
        'sells': balanced_sell_orders
    }
    
    # Resumo das alocações
    logger.info(f"Exposição balanceada COMPRA: R$ {exposicao_final_compra:.2f} de R$ {exposicao_por_lado:.2f} ({exposicao_final_compra/exposicao_por_lado*100:.2f}%)")
    logger.info(f"Exposição balanceada VENDA: R$ {exposicao_final_venda:.2f} de R$ {exposicao_por_lado:.2f} ({exposicao_final_venda/exposicao_por_lado*100:.2f}%)")
    logger.info(f"Total de ações compra: {len(balanced_buys)}")
    logger.info(f"Total de ações venda: {len(balanced_sells)}")
    logger.info(f"Ordens geradas: {len(balanced_buy_orders)} compra, {len(balanced_sell_orders)} venda")
    
    return orders

def get_stocks_list():
    """
    Retorna a lista de ações para monitorar a partir do arquivo ibra_stocks.json
    
    Returns:
        list: Lista de tickers com sufixo .SA
    """
    try:
        # Definir caminhos
        from config import DATA_DIR
        ibra_file = os.path.join(DATA_DIR, "ibra_stocks.json")
        
        # Verificar se existe arquivo com ações do IBRA
        if os.path.exists(ibra_file):
            with open(ibra_file, 'r', encoding='utf-8') as f:
                ibra_stocks = json.load(f)
            
            # Extrair tickers (usar yahoo_ticker se disponível)
            stocks = [stock['yahoo_ticker'] if stock.get('yahoo_ticker') else f"{stock['ticker']}.SA" 
                      for stock in ibra_stocks]
            
            # Limitar se configurado
            if CONFIG["market"].get("num_ibra_stocks"):
                num_stocks = CONFIG["market"]["num_ibra_stocks"]
                if isinstance(num_stocks, int) and num_stocks > 0 and num_stocks < len(stocks):
                    logger.info(f"Limitando para os {num_stocks} primeiros stocks do IBRA")
                    stocks = stocks[:num_stocks]
            
            logger.info(f"Carregados {len(stocks)} ativos do IBRA de {ibra_file}")
            return stocks
        else:
            # Fallback para lista default se arquivo não existir
            logger.warning(f"Arquivo {ibra_file} não encontrado, usando lista default")
            default_stocks = [
                'PETR4.SA', 'VALE3.SA', 'ITUB4.SA', 'BBDC4.SA', 'ABEV3.SA', 'B3SA3.SA', 
                'WEGE3.SA', 'RENT3.SA', 'BBAS3.SA', 'MGLU3.SA', 'RADL3.SA', 'SUZB3.SA',
                'JBSS3.SA', 'LREN3.SA', 'RAIL3.SA', 'GGBR4.SA', 'UGPA3.SA', 'BPAC11.SA',
                'VIVT3.SA', 'CMIG4.SA'
            ]
            logger.info(f"Usando lista default com {len(default_stocks)} ações")
            return default_stocks
    except Exception as e:
        logger.error(f"Erro ao carregar lista de ações: {e}")
        logger.info("Usando lista default")
        # Lista default caso ocorra algum erro
        default_stocks = [
            'PETR4.SA', 'VALE3.SA', 'ITUB4.SA', 'BBDC4.SA', 'ABEV3.SA', 'B3SA3.SA', 
            'WEGE3.SA', 'RENT3.SA', 'BBAS3.SA', 'MGLU3.SA'
        ]
        return default_stocks

def run_trading_system(exposicao_financeira_override=None):
    """
    Executa o sistema de trading FuzzyFajuto
    
    Args:
        exposicao_financeira_override (float, optional): Sobreescreve a exposição financeira configurada
        
    Returns:
        dict: Ordens geradas
    """
    try:
        logger.info("=== Iniciando sistema FuzzyFajuto ===")
        
        # Obter lista de ações
        symbols = get_stocks_list()
        logger.info(f"Processando {len(symbols)} ações")
        
        # Definir datas
        end_date = datetime.datetime.now()
        lookback_days = CONFIG["model"].get("lookback_days", 30)
        start_date = end_date - datetime.timedelta(days=lookback_days)
        logger.info(f"Período de análise: {start_date.date()} a {end_date.date()} ({lookback_days} dias)")
        
        # Baixar dados
        stock_data = get_stock_data(symbols, start_date, end_date)
        logger.info(f"Dados obtidos para {len(stock_data)} ações")
        
        # Calcular indicadores
        stock_data = calculate_indicators(stock_data)
        logger.info("Indicadores calculados")
        
        # Data mais recente disponível para todos os ativos
        last_date = end_date.date()
        for symbol, data in stock_data.items():
            if not data.empty:
                last_date = min(last_date, data.index[-1].date())
        
        logger.info(f"Usando data de referência: {last_date}")
        
        # Calcular scores 
        scores = calculate_fuzzy_fajuto_score(stock_data, pd.Timestamp(last_date))
        logger.info(f"Scores calculados para {len(scores)} ações")
        
        # Selecionar portfólio
        portfolio = select_portfolio(scores)
        logger.info(f"Portfólio selecionado com {len(portfolio['buys'])} compras e {len(portfolio['sells'])} vendas")
        
        # Gerar ordens
        orders = generate_tryd_orders(portfolio, exposicao_financeira_override)
        
        # Resumo final das ordens
        logger.info(f"Ordens geradas: {len(orders['buys'])} compra, {len(orders['sells'])} venda")
        
        return orders, portfolio
        
    except Exception as e:
        logger.error(f"Erro no sistema de trading: {e}")
        logger.error(traceback.format_exc())
        return {'buys': [], 'sells': []}, {}

def export_to_tryd_automate(orders, output_file='automate.xlsx'):
    """
    Exporta as ordens geradas para o formato do Tryd Automate
    
    Args:
        orders (dict): Ordens geradas pelo sistema
        output_file (str): Caminho do arquivo de saída
    """
    try:
        # Corrigir cabeçalho e colunas para seguir exatamente o padrão solicitado
        header = [
            'Papel', 'Cód. Cliente', 'Cond. Compra', 'Máx. Qtd. Compra', 'Qtd. Apar. Compra',
            'Qtd. Compra', 'Preço Compra', 'Preço Venda', 'Qtd. Venda', 'Qtd. Apar. Venda',
            'Máx. Qtd. Venda', 'Cond. Venda', 'Observação'
        ]
        try:
            wb = load_workbook(output_file)
            logger.info(f"Planilha existente carregada: {output_file}")
            if 'Sheet1' in wb.sheetnames:
                ws = wb['Sheet1']
                ws.delete_rows(2, ws.max_row)
            else:
                ws = wb.create_sheet('Sheet1')
                ws.append(header)
        except:
            logger.info(f"Criando nova planilha: {output_file}")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.append(header)
        client = "1234"
        today = datetime.datetime.now().strftime("%d/%m/%Y")
        for order in orders['buys']:
            symbol = order['stock']
            qty = order['qty']
            description = order.get('description', '')
            if order['price'] == 'MARKET':
                cond_compra = 'DIA'
                price_value = truncate_price(order.get('actual_price', order['exposure'] / order['qty']), 2)
                price = str(price_value).replace('.', ',')
            else:
                cond_compra = 'DIA'
                price_value = truncate_price(float(order['price']), 2)
                price = str(price_value).replace('.', ',')
            row_data = [
                symbol, client, cond_compra, qty, qty, qty, price,
                None, None, None, None, None, f"{today} - {description}"
            ]
            ws.append(row_data)
        for order in orders['sells']:
            symbol = order['stock']
            qty = order['qty']
            description = order.get('description', '')
            if order['price'] == 'MARKET':
                cond_venda = 'DIA'
                price_value = truncate_price(order.get('actual_price', order['exposure'] / order['qty']), 2)
                price = str(price_value).replace('.', ',')
            else:
                cond_venda = 'DIA'
                price_value = truncate_price(float(order['price']), 2)
                price = str(price_value).replace('.', ',')
            row_data = [
                symbol, client, None, None, None, None, None,
                price, qty, qty, qty, cond_venda, f"{today} - {description}"
            ]
            ws.append(row_data)
        wb.save(output_file)
        logger.info(f"Planilha salva com sucesso: {output_file}")
        
    except Exception as e:
        logger.error(f"Erro ao exportar para Excel: {e}")
        logger.error(traceback.format_exc())

def generate_risk_report(portfolio, orders, output_file=None):
    """
    Gera um relatório de risco baseado no portfólio e ordens
    
    Args:
        portfolio (dict): Portfólio selecionado
        orders (dict): Ordens geradas
        output_file (str): Caminho do arquivo de saída. Se None, será gerado automaticamente na pasta reports.
    """
    try:
        # Criar pasta reports se não existir
        from config import RESULTS_DIR
        reports_dir = os.path.join(RESULTS_DIR, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        # Gerar nome de arquivo com timestamp amigável se não for fornecido
        if output_file is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(reports_dir, f"risk_report_{timestamp}.txt")
        
        # Extrair informações necessárias
        compras = []
        vendas = []
        
        # Preparar dados para análise
        # Processar posições de compra
        buy_symbols = set()
        for order in orders['buys']:
            if order['type'] == 'MARKET':  # Usar apenas ordens de mercado para não duplicar
                buy_symbols.add(order['stock'])
                compras.append({
                    'ativo': order['stock'],
                    'preco': order['price'] if order['price'] != 'MARKET' else order['exposure'] / order['qty'],
                    'quantidade': order['qty'],
                    'exposicao': order['exposure'],
                    'score': order['score']
                })
        
        # Processar posições de venda
        sell_symbols = set()
        for order in orders['sells']:
            if order['type'] == 'MARKET':
                sell_symbols.add(order['stock'])
                vendas.append({
                    'ativo': order['stock'],
                    'preco': order['price'] if order['price'] != 'MARKET' else order['exposure'] / order['qty'],
                    'quantidade': order['qty'],
                    'exposicao': order['exposure'],
                    'score': order['score']
                })
        
        # Calcular métricas
        exposicao_compra = sum(item['exposicao'] for item in compras)
        exposicao_venda = sum(item['exposicao'] for item in vendas)
        exposicao_total = exposicao_compra + exposicao_venda
        exposicao_net = exposicao_compra - exposicao_venda
        
        # Calcular orçamento total baseado na configuração
        orcamento_total = CONFIG["model"]["exposicao_financeira_total"]
        orcamento_por_lado = orcamento_total / 2
        
        # Uso do orçamento
        uso_orcamento_compra = (exposicao_compra / orcamento_por_lado) * 100
        uso_orcamento_venda = (exposicao_venda / orcamento_por_lado) * 100
        
        # Calcular índice de concentração HHI
        hhi_compra = sum([(item['exposicao'] / exposicao_compra * 100) ** 2 for item in compras]) if exposicao_compra > 0 else 0
        hhi_venda = sum([(item['exposicao'] / exposicao_venda * 100) ** 2 for item in vendas]) if exposicao_venda > 0 else 0
        
        # Identificar maiores posições
        maior_posicao_compra = max(compras, key=lambda x: x['exposicao']) if compras else None
        maior_posicao_venda = max(vendas, key=lambda x: x['exposicao']) if vendas else None
        
        # Gerar relatório
        current_time = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        report_lines = [
            "=" * 80,
            f"RELATÓRIO DE RISCO FUZZYFAJUTO - {current_time}",
            "=" * 80,
            "",
            "1. PARÂMETROS DO MODELO",
            "-" * 50,
            f"Exposição financeira total: R$ {orcamento_total:.2f}",
            f"Máximo de posições por lado: {CONFIG['model']['max_position_per_side']}",
            f"Dias de histórico utilizados: {CONFIG['model'].get('lookback_days', 30)}",
            "",
            "2. RESUMO DE POSIÇÕES",
            "-" * 50,
            f"Número de posições de compra: {len(compras)}",
            f"Número de posições de venda: {len(vendas)}",
            f"Exposição total de compra: R$ {exposicao_compra:.2f}",
            f"Exposição total de venda: R$ {exposicao_venda:.2f}",
            f"Exposição líquida (net): R$ {exposicao_net:.2f}",
            f"Uso do orçamento de compra: {uso_orcamento_compra:.2f}%",
            f"Uso do orçamento de venda: {uso_orcamento_venda:.2f}%",
            "",
            "3. POSIÇÕES DE COMPRA",
            "-" * 50,
            f"{'Ativo':<8} {'Preço':<10} {'Quantidade':<12} {'Exposição (R$)':<15} {'% do Total':<12} {'Score':<10}",
            "-" * 80,
        ]
        
        # Adicionar posições de compra
        for item in sorted(compras, key=lambda x: x['exposicao'], reverse=True):
            percentual = (item['exposicao'] / orcamento_total) * 100
            price_str = f"R$ {item['preco']:.2f}" if item['preco'] != 'MARKET' else 'MERCADO'
            report_lines.append(f"{item['ativo']:<8} {price_str:<10} {item['quantidade']:<12} R$ {item['exposicao']:.2f}{'':<5} {percentual:<6.2f}{'%':<6} {item['score']:<10.2f}")
        
        report_lines.extend([
            "",
            "4. POSIÇÕES DE VENDA",
            "-" * 50,
            f"{'Ativo':<8} {'Preço':<10} {'Quantidade':<12} {'Exposição (R$)':<15} {'% do Total':<12} {'Score':<10}",
            "-" * 80,
        ])
        
        # Adicionar posições de venda
        for item in sorted(vendas, key=lambda x: x['exposicao'], reverse=True):
            percentual = (item['exposicao'] / orcamento_total) * 100
            price_str = f"R$ {item['preco']:.2f}" if item['preco'] != 'MARKET' else 'MERCADO'
            report_lines.append(f"{item['ativo']:<8} {price_str:<10} {item['quantidade']:<12} R$ {item['exposicao']:.2f}{'':<5} {percentual:<6.2f}{'%':<6} {item['score']:<10.2f}")
        
        report_lines.extend([
            "",
            "5. ANÁLISE DE RISCO",
            "-" * 50,
        ])
        
        # Adicionar análise de risco
        if maior_posicao_compra:
            perc_maior_compra = (maior_posicao_compra['exposicao'] / orcamento_total) * 100
            report_lines.append(f"Maior posição de compra: {maior_posicao_compra['ativo']} - R$ {maior_posicao_compra['exposicao']:.2f} ({perc_maior_compra:.2f}% do orçamento)")
        
        if maior_posicao_venda:
            perc_maior_venda = (maior_posicao_venda['exposicao'] / orcamento_total) * 100
            report_lines.append(f"Maior posição de venda: {maior_posicao_venda['ativo']} - R$ {maior_posicao_venda['exposicao']:.2f} ({perc_maior_venda:.2f}% do orçamento)")
        
        report_lines.append(f"Índice de concentração HHI compra: {hhi_compra:.2f}/10000 (menor é melhor)")
        report_lines.append(f"Índice de concentração HHI venda: {hhi_venda:.2f}/10000 (menor é melhor)")
        
        # Escrever relatório no arquivo
        with open(output_file, 'w', encoding='utf-8') as file:
            file.write('\n'.join(report_lines))
        
        logger.info(f"Relatório de risco salvo em: {output_file}")
        
        # Também exibir no console
        print('\n'.join(report_lines))
        
        return output_file
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório de risco: {e}")
        logger.error(traceback.format_exc())
        return None

if __name__ == "__main__":
    # Executar sistema
    orders, portfolio = run_trading_system()
    
    # Exportar ordens para Tryd
    export_to_tryd_automate(orders)
    
    # Gerar relatório de risco
    generate_risk_report(portfolio, orders)
    
    logger.info("Processamento concluído!") 